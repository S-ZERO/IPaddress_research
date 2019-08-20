import openpyxl
import pandas as pd

#取得する最初のセルと最後のセルを取得→データフレームを返すまでやる関数
def get_cell_df(def_title,maxrow,sheet):
    #-------------文字列操作で無理やり最初・最後のセル番地を取得-------------
    def_title_str = str(def_title)
    first_cell_tmp =  def_title_str.strip("<>").split(".")
    first_cell = first_cell_tmp[1]
    second_cell = first_cell.strip("1")+ "2"
    last_cell = first_cell.strip("1")+ str(maxrow)
    #-------------文字列操作で無理やり最初・最後のセル番地を取得-------------

    #pd.DataFrameに投げ込む用のセル一覧作成用のリスト
    cells = []

    #cell一覧を取得して、cellsに一つずつappend
    for target_field in sheet[second_cell:last_cell]:
        for cell in target_field:
            cells.append(cell.value)

    #cellsをDataFrame化
    df_cells = pd.DataFrame(cells, columns = [def_title.value])

    #返り血にdef_titleを含んだのは、drop_duplicateを実行するときに、データフレームのcolumnsが必要になるから
    return df_cells, def_title

#get_cell_dfで出力したデータフレームを結合する関数(無駄だったかも)
def concat_df(target_addr, info):
    #送信元とその内外情報　送信先とその内外情報をそれぞれ結合
    shape_addr = pd.concat([target_addr, info], axis=1)
    return shape_addr
 

print("対象となるエクセルのファイルパス群を入力")
target_xlsx_path = input(">>")

xlsx = []

with open(target_xlsx_path, encoding="utf-8") as f:
    for rows in f:
        row = rows.rstrip("\n")
        xlsx.append(row)

for target_xlsx in xlsx:
    target_book = openpyxl.load_workbook(target_xlsx)
    target_book_sheet = target_book["全体"]
    #max_columnは何項目あるかを表していてを抜き出す項目分繰り返すために使用
    target_book_maxcol = target_book_sheet.max_column
    target_book_maxrow = target_book_sheet.max_row
    print("--------------------------------------")
    print("検索する項目数："+str(target_book_maxcol))
    print("出力する情報の件数:"+str(target_book_maxrow))
    print("--------------------------------------")
    #sheet名.iter_rowsは選んだ範囲のセルをタプル二次元配列で返すから
    for titles in target_book_sheet.iter_rows(min_row=1,min_col=1,max_row=1,max_col=target_book_maxcol):
        #.iter_rowsで得た結果でfor分を回していく
        for title in titles:
            #ほしいもの自体は取得できたから、あとはdf化して社内列もdf化して、社内列が[社外]のやつだけ抜き出す
            if title.value=="SourceAddr":
                SourceAddr = get_cell_df(title, target_book_maxrow, target_book_sheet)
                #print(SourceAddr)

            elif  title.value == "送信元\n(内外)":
                SourceAddr_info = get_cell_df(title, target_book_maxrow, target_book_sheet)
                #print(SourceAddr_info)
                
            elif title.value == "DestAddr":
                DestAddr = get_cell_df(title, target_book_maxrow, target_book_sheet)
                #print(DestAddr)

            elif  title.value == "宛先\n(内外)":
                DestAddr_info = get_cell_df(title, target_book_maxrow, target_book_sheet)
                #print(DestAddr_info)

            elif title.value == ":server":
                server = get_cell_df(title, target_book_maxrow, target_book_sheet)
                #server重複排除後
                server_after_tyoufuku = server[0].drop_duplicates([str(server[1].value)]) 
                print("----------------------------------")
                print("重複排除前")
                print(server[0])
                print("----------------------------------")    
    
                print("----------------------------------")
                print("重複排除結果")
                print(server_after_tyoufuku)
                print("----------------------------------")

    #送信先・送信元の二つに関しては社内IPについては調べる必要がないため、社内IPであれば消去する
    #消去する前段階として社内IPなのか社外IPなのかを示すために、社内社外情報と結合する→社内データフレームがついていたらそれをトリガーにしてー削除する
    #また、serverは問答無用で調査対象だから、社内か社外かとかどうでもいい→社内社外情報と結合する必要がないからしない
    
    #送信元IP重複排除前
    SourceAddr_before_tyoufuku = concat_df(SourceAddr[0], SourceAddr_info[0])
    #送信先IP重複排除前
    DestAddr_before_tyoufuku = concat_df(DestAddr[0], DestAddr_info[0])

    #送信元IP重複排除後
    SourceAddr_after_tyoufuku = SourceAddr_before_tyoufuku.drop_duplicates([str(SourceAddr[1].value), str(SourceAddr_info[1].value)])
    #送信先IP重複排除後
    DestAddr_after_tyoufuku = DestAddr_before_tyoufuku.drop_duplicates([str(DestAddr[1].value), str(DestAddr_info[1].value)])
    
    print("----------------------------------")
    print("重複排除前")
    print(SourceAddr_before_tyoufuku)
    print("----------------------------------")
    
    print("----------------------------------")
    print("重複排除結果")
    print(SourceAddr_after_tyoufuku)
    print("----------------------------------") 
    
    print("----------------------------------")
    print("重複排除前")
    print(DestAddr_before_tyoufuku)
    print("----------------------------------")    
    
    print("----------------------------------")
    print("重複排除結果")
    print(DestAddr_after_tyoufuku)
    print("----------------------------------")
   


