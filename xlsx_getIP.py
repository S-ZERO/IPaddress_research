import openpyxl

#取得する最初のセルと最後のセルを取得
def get_fcell(def_title,maxrow):
    def_title = str(def_title)
    first_cell_tmp =  def_title.strip("<>").split(".")
    first_cell = first_cell_tmp[1]
    print(first_cell)

    last_cell = first_cell.strip("1")+ str(maxrow)

    return first_cell,last_cell

print("対象となるエクセルのファイルパス群を入力")
target_xlsx_path = input(">>")

xlsx = []

with open(target_xlsx_path, encoding="utf-8") as f:
    for rows in f:
        row = rows.rstrip("\n")
        xlsx.append(row)

for target_xlsx in xlsx:
    target_book = openpyxl.load_workbook(target_xlsx)
    target_book_sheet = target_book["全体"]
    target_book_maxcol = target_book_sheet.max_column
    target_book_maxrow = target_book_sheet.max_row
    print(target_book_maxcol)
    #sheet名.iter_rowsは選んだ範囲のセルをタプル二次元配列で返すから
    for titles in target_book_sheet.iter_rows(min_row=1,min_col=1,max_row=1,max_col=target_book_maxcol):
        #.iter_rowsで得た結果でfor分を回していく
        for title in titles:
            #ほしいもの自体は取得できたから、あとはdf化して社内列もdf化して、社内列が[社外]のやつだけ抜き出す
            if title.value=="SourceAddr":
                cell_range = get_fcell(title,target_book_maxrow)
                print(title.value)
                for rows in target_book_sheet[cell_range[0]:cell_range[1]]:
                    for cell in rows:
                        print(cell.value)

            elif title.value == "DestAddr":
                get_fcell(title,target_book_maxrow)
                cell_range = get_fcell(title,target_book_maxrow)
                print(title.value)
                for rows in target_book_sheet[cell_range[0]:cell_range[1]]:
                    for cell in rows:
                        print(cell.value)

            elif title.value == ":server":
                cell_range = get_fcell(title,target_book_maxrow)
                print(title.value)
                for rows in target_book_sheet[cell_range[0]:cell_range[1]]:
                    for cell in rows:
                        print(cell.value)


