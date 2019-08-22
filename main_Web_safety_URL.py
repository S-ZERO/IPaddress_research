from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException

from time import sleep
import os, argparse, csv

from bs4 import BeautifulSoup

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from requests.exceptions import Timeout

from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException

import aguse_reserach_module as aguse
import mxtoolbox_blacklist_URL_module as mxtoolbox
import owner_search_URL_module as ownerURL
import xforce_exchange_URL_module as xforce
import xforce_exchange_login_module as xforce_login

import pandas as pd
import openpyxl
from openpyxl import load_workbook

search = []

aguse_results = []
mxtoolbox_results = []
owner_URL_results = []

xforce_category = []
xforce_time = []
xforce_risk = []
xforce_URL = []

def all_result_writer(write_result_list, col_num, target_wb):
    headers = ["IP", "aguse", "mxtoolbox", "owner"]
    fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor='228B22', bgColor='228B22')
    font = openpyxl.styles.fonts.Font(color = 'FFFFFF', bold = True )
    border_line = openpyxl.styles.borders.Side(style='thin', color = '000000')
    border = openpyxl.styles.borders.Border(top=border_line, bottom=border_line, left=border_line, right=border_line)
    print(fill)
    if not "all_result" in target_wb.get_sheet_names():
        all_result_sheet = target_wb.create_sheet(title = "all_result")
        for col, header in enumerate(headers):
            all_result_sheet.cell(row = 1, column = col+1, value = header).fill = fill
            all_result_sheet.cell(row = 1, column = col+1).font = font
            all_result_sheet.cell(row = 1, column = col+1).border = border
    else:
        all_result_sheet = target_wb.get_sheet_by_name("all_result")

    for num, write_value in enumerate(write_result_list):
        all_result_sheet.cell(row = num+2, column = col_num+1, value = write_value)
        all_result_sheet.cell(row = num+2, column = col_num+1).border = border
    
    wb.save(r"C:\PF_IP_searcher\result\result.xlsx")

driver = webdriver.Chrome(r'C:\PF_IP_searcher\chromedriver.exe') 

print("読み込ませるファイルを選択")
target_file = input(">> ")
with open(target_file, encoding='utf-8') as f:
    for rows in f:
        row = rows.rstrip('\n\n')
        search.append(row)

for url in search:
    #------aguseの処理--------------------------------
    for roop in range(0,5):
        aguse_result = aguse.main_move(url,driver)
        if aguse_result == "safe" or aguse_result == "caution":
            break
    aguse_results.append(aguse_result)
    print(aguse_results)
    
    #------mxtoolboxの処理--------------------------------
    for roop in range(0,6):#ループが6まである理由は、aguseの妨害処理に1ループ無駄にするかもしれないから
       mxtoolbox_result = mxtoolbox.main_move(url,driver)
       if not mxtoolbox_result == None and int(mxtoolbox_result[1]) < 5:
           break
    mxtoolbox_results.append(mxtoolbox_result[0])
    print(mxtoolbox_results)
    
    #------ownerURLの処理--------------------------------
    for roop in range(0,5):
       owner_URL_result = ownerURL.main_move(url,driver)
       if not owner_URL_result == None and owner_URL_result != "Web接続制限":
           break
    owner_URL_results.append(owner_URL_result)
    print(owner_URL_results)

#---xforce exchangeの処理------------------------
xforce_login.login(driver)
for url in search:
    for roop in range(0,5):
        xforce_result = xforce.main_move(url, driver)
        
        if xforce_result == "Web接続制限":
            xforce_login.login(driver)
                    
        elif not xforce_result == None:
            break
            
    xforce_results_category = xforce_result[0]
    for category in xforce_results_category:
        xforce_category.append(category)
    xforce_results_time = xforce_result[1]
    for time in xforce_results_time:
        xforce_time.append(time)
    xforce_results_risk = xforce_result[2]
    for risk in xforce_results_risk:
        xforce_risk.append(risk)
    xforce_results_URL = xforce_result[3]
    for IP in xforce_results_URL:
        xforce_URL.append(url)
    print(xforce_category)
    print(xforce_time)
    print(xforce_risk)
    print(xforce_URL)

wb = openpyxl.Workbook()
all_results_list = [search, aguse_results, mxtoolbox_results, owner_IP_results]
for num, results_list in enumerate(all_results_list):
    all_result_writer(results_list, num, wb)


# #--------以下結果の出力--------------------------------------------------------------------------------
# #検索したURLアドレスと全部の結果をデータフレームにする
# df_searchURL = pd.DataFrame(search, columns=["URL"])
# df_aguse_results = pd.DataFrame(aguse_results, columns=["aguse"])
# df_mxtoolbox_results = pd.DataFrame(mxtoolbox_results, columns=["mxtoolbox"])
# df_owner_URL_results = pd.DataFrame(owner_URL_results, columns=["ページタイトル"])

# #データフレームにした結果を結合
# df_all_results_tmp = pd.concat([df_searchURL, df_aguse_results, df_mxtoolbox_results, df_owner_URL_results], axis=1)

# df_all_results = df_all_results_tmp.set_index(['URL', 'aguse', 'mxtoolbox', 'ページタイトル'])

# #--------xforceの結果出力------------------------------------------------
# df_xforceIP = pd.DataFrame(xforce_URL, columns=["URL"])
# df_xforce_category = pd.DataFrame(xforce_category, columns=["カテゴリー"])
# df_xforce_time = pd.DataFrame(xforce_time, columns=["タイムライン"])
# df_xforce_risk = pd.DataFrame(xforce_risk, columns=["リスク"])

# df_xforce_all_results_tmp = pd.concat([df_xforceIP, df_xforce_category, df_xforce_time, df_xforce_risk], axis=1)

# df_xforce_all_results = df_xforce_all_results_tmp.set_index(["URL", "カテゴリー", "タイムライン", "リスク"])


# #----------test.xlsxを開いて、結果シートとxforce結果シートに結果を追記-------------------------------------------
# writer = pd.ExcelWriter(r'C:\PF_IP_searcher\result\restult_template.xlsx')
# writer.book = openpyxl.load_workbook(r'C:\PF_IP_searcher\result\restult.xlsx')
# df_all_results.to_excel(writer, sheet_name="結果")
# df_xforce_all_results.to_excel(writer, sheet_name="xforce結果")

print("完了")
driver.close()